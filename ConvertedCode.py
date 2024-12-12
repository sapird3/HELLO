# Project for Adam Calvart
# BY DANA SAPIR
# 07/09/2024

# convert the VBA code (from the Prochain Formatter Excel) into Python
# involves using the openpyxl library for working with Excel files and replicating the functionality of the VBA macros in Python

import openpyxl
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

# Globals
g_strTitle = "Your Application Title"
g_strVersion = "1.0"
g_strCopyright = "Your Company"
ARTScrumTeam = [
    "Foundation - Astra NPS Developers", "Foundation - Atlas", "Foundation - Bart", "Foundation - Boosters",
    # Add the rest of your teams here...
]

def is_file_found(str_path):
    return os.path.isfile(str_path)

def worksheet_exists(workbook, worksheet_name):
    return worksheet_name in workbook.sheetnames

def standardize_string(s):
    return s.strip().replace(" ", "").replace("_", "").upper()

def is_cyber_struc_export(ws):
    return (
        standardize_string(ws['A1'].value) == standardize_string("Index") and
        standardize_string(ws['B1'].value) == standardize_string("Key") and
        standardize_string(ws['C1'].value) == standardize_string("Successors") and
        standardize_string(ws['D1'].value) == standardize_string("Predecessors") and
        standardize_string(ws['E1'].value) == standardize_string("Summary") and
        standardize_string(ws['F1'].value) == standardize_string("Issue Type") and
        standardize_string(ws['G1'].value) == standardize_string("Status") and
        standardize_string(ws['H1'].value) == standardize_string("CommitmentLevel") and
        standardize_string(ws['I1'].value) == standardize_string("Sprint") and
        standardize_string(ws['J1'].value) == standardize_string("Scrum Teams") and
        standardize_string(ws['K1'].value) == standardize_string("ART Scrum Team") and
        standardize_string(ws['L1'].value) == standardize_string("Parent Feature") and
        standardize_string(ws['M1'].value) == standardize_string("Astra_NPS_Developers_TG") and
        standardize_string(ws['N1'].value) == standardize_string("Astra_NPS_Developers_TC") and
        standardize_string(ws['O1'].value) == standardize_string("Astra_NPS_Developers_FD") and
        standardize_string(ws['P1'].value) == standardize_string("Astra_NPS_Developers_LRD") and
        standardize_string(ws['Q1'].value) == standardize_string("Astra_NPS_Developers_FD_Remaining")
    )

def is_milestone(ws):
    return (
        standardize_string(ws['A1'].value) == standardize_string("Selector ( Text10 )") and
        standardize_string(ws['B1'].value) == standardize_string("Unique ID") and
        standardize_string(ws['C1'].value) == standardize_string("Unique ID Successors") and
        standardize_string(ws['D1'].value) == standardize_string("Structure ID ( Text11 )") and
        standardize_string(ws['E1'].value) == standardize_string("Parent Feature ( Text12 )") and
        standardize_string(ws['F1'].value) == standardize_string("Key ( Text13 )") and
        standardize_string(ws['G1'].value) == standardize_string("Name") and
        standardize_string(ws['H1'].value) == standardize_string("Issue Type ( Text14 )") and
        standardize_string(ws['I1'].value) == standardize_string("Status ( Text15 )") and
        standardize_string(ws['J1'].value) == standardize_string("ART ( Text16 )") and
        standardize_string(ws['K1'].value) == standardize_string("Resource Names") and
        standardize_string(ws['L1'].value) == standardize_string("Task Type ( Text17 )") and
        standardize_string(ws['M1'].value) == standardize_string("Source ( Text18 )") and
        standardize_string(ws['N1'].value) == standardize_string("TG ( Text19 )") and
        standardize_string(ws['O1'].value) == standardize_string("TC ( Text20 )") and
        standardize_string(ws['P1'].value) == standardize_string("FD ( Duration )") and
        standardize_string(ws['Q1'].value) == standardize_string("LRD ( Duration1 )") and
        standardize_string(ws['R1'].value) == standardize_string("FDR ( Text21 )") and
        standardize_string(ws['S1'].value) == standardize_string("% Complete")
    )

def log_entry(ws, n_row, str_event, str_data):
    ws.cell(row=n_row, column=1).value = datetime.now()
    ws.cell(row=n_row, column=2).value = str_event
    ws.cell(row=n_row, column=3).value = str_data
    return n_row + 1

def find_art(str_scrum_team):
    for team in ARTScrumTeam:
        if str_scrum_team.upper() in team.upper():
            return team.split(" - ")[0].strip()
    return "UNKNOWN"

def prochain_formatter_main():
    # Initialize some variables
    dt_processing_start_time = datetime.now()
    
    # File selection and workbook loading
    str_path_jira = "path_to_your_jira_file.xlsx"  # Replace with actual file path
    wb_jira = openpyxl.load_workbook(str_path_jira)
    ws_data_jira = wb_jira.active
    
    if not is_cyber_struc_export(ws_data_jira):
        print(f"{str_path_jira} spreadsheet is not formatted as expected. Exiting the application.")
        return
    
    # Create Project Data File
    str_date_time_stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    str_path_project = f"{str_date_time_stamp}_Project_Import.xlsx"
    wb_project = openpyxl.Workbook()
    ws_data_project = wb_project.active
    ws_data_project.title = "Project"
    
    # Add Header
    headers = [
        "Text10", "Unique ID", "Unique ID Successors", "Unique ID Predecessors", "Text11", "Text12", "Text13",
        "Name", "Text14", "Text15", "Text16", "Resource Names", "Text17", "Text18", "Text19", "Text20", "Duration",
        "Duration1", "Text21", "% Complete"
    ]
    for col_num, header in enumerate(headers, 1):
        ws_data_project.cell(row=1, column=col_num).value = header
    
    # Add Summary and Log Sheets
    ws_summary_project = wb_project.create_sheet(title="Summary")
    ws_log_project = wb_project.create_sheet(title="Log")
    
    # Add Header to Log Row sheet
    n_row_log = 1
    log_headers = ["DateTimeStamp", "Event", "Data"]
    for col_num, header in enumerate(log_headers, 1):
        ws_log_project.cell(row=n_row_log, column=col_num).value = header
    n_row_log += 1
    
    # Log Entries
    log_entry(ws_log_project, n_row_log, "Start:", f"{g_strTitle}, {g_strVersion}")
    log_entry(ws_log_project, n_row_log, "Input File:", str_path_jira)
    log_entry(ws_log_project, n_row_log, "Output File:", str_path_project)
    
    # Process Jira Data
    n_row_project = 2
    for row in ws_data_jira.iter_rows(min_row=2, values_only=True):
        # Process each row
        pass
    
    # Save the project workbook
    wb_project.save(str_path_project)
    
    # Processing complete
    dt_processing_end_time = datetime.now()
    log_entry(ws_log_project, n_row_log, "Processing Complete:", f"{(dt_processing_end_time - dt_processing_start_time).total_seconds()} seconds")

    # Save the log workbook
    wb_project.save(str_path_project)

# Run the main function
prochain_formatter_main()
