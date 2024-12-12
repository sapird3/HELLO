import xlwings as xw
import win32com.client
import time
import openpyxl
import json
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.styles.colors import BLACK, WHITE
from openpyxl.utils import get_column_letter
import shutil
import psutil
import sys
import os

import tkinter as tk
from tkinter import filedialog

class SentinelScheduleUpdater:
    def __init__(self, sentinelStructFilename, formatterFilename, projectFilename, baselineScheduleFilename, pathToScheduleUpdateSummary, debugFlag):
        print('Welcome to the automated schedule updating tool made by Barret Daniels')
        self.sentinelStructFilename = str(sentinelStructFilename)
        self.formatterFilename = str(formatterFilename)
        self.projectFilename = str(projectFilename)
        self.data_dict = {}
        self.tasksNotInStruct = {}
        self.com_max_retries = 20
        self.com_retry_delay = 1.2
        self.debugFlag = debugFlag #used to be more verbose and export intermediate JSON files for troubleshooting issues
        self.badTicketFixTicketsMarkedComplete = []
        self.baselineScheduleFilename = str(baselineScheduleFilename)
        #self.formattedExcelForImport = os.path.dirname(self.sentinelStructFilename) + "/Project_Import.xlsx"
        formattedExcelForImport = filedialog.askopenfilename(title="Select the latest Excel file formated for import", filetypes=(("Excel files", "*.xlsx"),  ("all files", "*.*")))
        self.formattedExcelForImport = formattedExcelForImport.replace('/', '\\')
        
        self.pathToScheduleUpdateSummary = str(pathToScheduleUpdateSummary)
        self.missingScrumTeamAssignment = []
        self.missingSuccessor = []
        self.newScrumTeamTickets = []
        
    def errorCheckStructureOutput(self):
        #Load workbook
        errorFoundInStruct = False
        print("This tool cannot presently detect duplicates, but JIRA structure can.  Recommend opening the structure and looking at the bottom where it will indicate the number of rows with duplicates.  You can press filter and then address the duplicates directly there and re-export.  Your export should contain no duplicates to avoid double counting work.")
        try:
            # Load the workbook
            wb = openpyxl.load_workbook(self.sentinelStructFilename)
            
            # Select the active sheet
            sheet = wb.active
        except Exception as e:
            print(f"Error processing Excel file: {e}")
            
        #Check for user stories that have only an ART set without a scrum
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Get the text in column K
            column_k_text = str(row[10])  # Column K is at index 10 (0-based index)
            column_f_text = str(row[5])

            # Check if the text in column K is not blank and doesn't contain a '-'
            if column_k_text.strip() != '' and '-' not in column_k_text and 'User Story' in column_f_text:
                # Print the text in column B followed by the text in column E
                #print(f"Ticket: {row[1]} {row[4]} has an ART set but no scrum team.  Please update the ART Scrum Team field in JIRA and re-export the structure.")  # Column B is at index 1, Column E is at index 4 (0-based index)
                self.missingScrumTeamAssignment.append({row[1] + " " + row[4]})
                errorFoundInStruct = True
        
        if len(self.missingScrumTeamAssignment) !=0:
            print ("Tickets missing Scrum Team Assignment in ART Scrum Field: ")
            for x in range(len(self.missingScrumTeamAssignment)):
                print (self.missingScrumTeamAssignment[x])
            print()
        
        #check for features without successors i.e making them endpoints in the schedule
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Get the text in column F
            column_f_text = str(row[5])  # Column F is at index 5 (0-based index)
            column_e_text = str(row[4]).lower()

            # Check if the word "Hello" is in column F
            if 'Feature' in column_f_text and 'prochain' not in column_e_text and 'milestone' not in column_e_text:
                # Check if column C is blank
                if not row[2]:  # Column C is at index 2 (0-based index)
                    # Print the text in column B and column E
                    #print(f"Ticket: {row[1]} {row[4]} is a feature and it has no successor.  Each feature must be a predecessor to another feature or a predecessor to a project milestone.  Please make a successor link in JIRA and re-export the structure")  # Column B is at index 1, Column E is at index 4 (0-based index)
                    self.missingSuccessor.append({row[1] + " " + row[4]})
                    errorFoundInStruct = True
        
        if len(self.missingSuccessor) !=0:
            print ("Features missing successor relationship: ")
            for x in range(len(self.missingSuccessor)):
                print (self.missingSuccessor[x])
            print ()
        
        #Check for missing scrum columns.  Note this is informative only and does not enforce a stopping of this tool.
        # Initialize an empty array to store the processed column headers
        processed_columns = []
        missing_scrums = []

        # Iterate through columns M to the end, skipping ahead 5 each time
        for col_index in range(13, sheet.max_column + 1, 5):
            # Get the text in row 1 of the current column
            column_header = sheet.cell(row=1, column=col_index).value

            # Remove the last 3 characters from the column header
            processed_header = str(column_header[:-3]).lower().replace("_"," ")

            # Add the processed header to the array
            processed_columns.append(processed_header)
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Get the text in column K
            column_k_text = str(row[10])  # Column K is at index 10 (0-based index)
            column_g_text = str(row[6])

            # Check if the text in column K is not blank and doesn't contain a '-'
            if column_k_text.strip() != '' and '-' in column_k_text:
                if '-' in column_k_text:
                    parts = column_k_text.split(' - ')
                    # Print the text in column B followed by the text in column E
                    if parts[1].lower() not in processed_columns and column_g_text != "Done" and column_g_text != "Cancelled":
                        #print(f"Scrum: {parts[1].lower()} set in Ticket: {row[1]} {row[4]} is not represented in the columns of the structure.  This means these points will not be counted in the resulting schedule.  It is recommended you fix this, but this error will not halt the schedule update.")
                        self.newScrumTeamTickets.append({row[1] + " " + parts[1].lower()})
                        if parts[1].lower() not in missing_scrums:
                            missing_scrums.append(parts[1].lower())
        
        # print("Scrums found in Structure Columns Export:")
        # print(processed_columns)
        if len(missing_scrums) !=0:
            print("Scrums missing from Structure Columns:")
            print(missing_scrums)
            print("In the following tickets: ")
            for x in range(len(self.newScrumTeamTickets)):
                print (self.newScrumTeamTickets[x])
            print ()
        
        if errorFoundInStruct:
            print("Schedule update cannot be run until the above issues in the JIRA structure are corrected in JIRA and a new export is generated without errors")
            sys.exit(1)
        else:
            print("CONGRATS!!  No issues found with JIRA structure export.  Proceeding to perform schedule update.")
    
    """
    def formatTheJIRAExport(self):
        print("Standby: Formatting the JIRA Export")
        excel_app = xw.App(visible=False)  # To run Excel in the background without displaying it
        workbook = excel_app.books.open(self.formatterFilename)

        macro_name = "ProchainFormatterMain"
        excel_app = xw.apps.active  # If Excel is already open
        excel_app.api.Run(macro_name, self.sentinelStructFilename)
        workbook.close()
        # Quit Excel
        excel_app.quit()
        print("Finsihed: Formatting the JIRA Export")
    """
    
    def buildDictionaryFromFormatterExport(self):
        print("Standby: Building Python Dictionary from Formatted Excel")
        try:
            workbook = openpyxl.load_workbook(self.formattedExcelForImport)
            sheet = workbook.active

            # Assuming the data starts from row 2
            for row in sheet.iter_rows(min_row=2, values_only=True):
                key = row[0]  # Value from column A
                formatterAssignedID = str(row[1]) or ""
                uniqueIDSuccessors = str(row[2]) or ""
                if uniqueIDSuccessors == "None":
                    uniqueIDSuccessors = ""
                uniqueIDPredecessors = str(row[3]) or ""
                if uniqueIDPredecessors == "None":
                    uniqueIDPredecessors = ""
                textEleven = row[4] or ""
                textTwelve = row[5] or ""
                textThirteen = row[6] or ""
                taskName =  row[7] or ""
                textFourteen = row[8] or ""
                textFifteen = row[9] or ""
                textSixteen = row[10] or ""
                resourceNames = row[11] or ""
                textSeventeen = row[12] or ""
                textEighteen = row[13] or ""
                textNinteen = str(row[14]) or ""
                textTwenty = str(row[15]) or ""
                focused_duration = str(row[16]) or ""
                if ("hr" not in focused_duration and focused_duration != ""):
                    focused_duration = focused_duration + " days"
                low_risk_duration = str(row[17]) or ""
                if ("hr" not in low_risk_duration and low_risk_duration != ""):
                    low_risk_duration = low_risk_duration + " days"
                remainingDuration = str(row[18]) or ""
                if ("hr" not in remainingDuration and remainingDuration != ""):
                    remainingDuration = remainingDuration + " days"
                percent_complete = str((row[19] * 100)) or ""
                tempNumber = float(percent_complete)
                roundedNumber = round(tempNumber, 2)
                percent_complete = str(roundedNumber)
                
                # Build the inner dictionary
                inner_dict = {
                    "IsProjectBuffer": "No",
                    "ProjectTaskNoStruct": "No",
                    "NewTask": "Yes",
                    "Unique ID": "",
                    "Formatter Assigned ID": formatterAssignedID,
                    "Unique ID Successors": uniqueIDSuccessors,
                    "Unique ID Predecessors": uniqueIDPredecessors,
                    "Text11": textEleven,
                    "Text12": textTwelve,
                    "Text13": textThirteen,
                    "Name": taskName,
                    "Text14": textFourteen,
                    "Text15": textFifteen,
                    "Text16": textSixteen,
                    "Resource Names": resourceNames,
                    "Text17": textSeventeen,
                    "Text18": textEighteen,
                    "Text19": textNinteen,
                    "Text20": textTwenty,
                    "Duration": focused_duration,
                    "Low-Risk Dur": low_risk_duration,
                    "Text21": remainingDuration,
                    "percentcomplete": percent_complete
                }

                # Add the inner dictionary to the main dictionary
                self.data_dict[key] = inner_dict
            workbook.save(self.formattedExcelForImport)
            workbook.close()

        except Exception as e:
            print(f"Error reading the Excel file: {e}")
        print("Finished: Building Python Dictionary from Formatted Excel")
    
    
    def deleteProjectTasksNoLongerInStruct(self):
        print("Standby: Finding and deleting MS project tasks that no longer exist in the structure")
        print("NOTE: This is slow on purpose because com32 is slow and error prone.  Please be patient.")
        # Create an instance of Microsoft Project
        try:
            app = win32com.client.Dispatch("MSProject.Application")
        except Exception as e:
            print(f"Error: {str(e)}")

        # Open the MPP file
        mpp_file_path = self.projectFilename
        app.FileOpen(mpp_file_path)
        time.sleep(5)
        # Get a reference to the project
        # Check if there is an active project
        if app.ActiveProject is not None:
            project = app.ActiveProject
            # Now you can work with the active project
        else:
            print("No active project found.")
        
        # Get a reference to the tasks in the project
        tasks = project.Tasks
        
        for task in tasks:
            taskTextTen = ""
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    taskTextTen = str(task.Text10)

                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
                        
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
            if taskTextTen != "":
                for retry_count in range(self.com_max_retries):
                    if taskTextTen not in self.data_dict:
                        try:
                            
                            # Open the Microsoft Project file
                            if taskTextTen[:2] != "PB": #necessary to ensure it doesn't delete project buffers
                                if self.debugFlag:
                                    print("Deleting following task no longer in the structure: " + taskTextTen)
                                task.Delete()

                            # If the operation succeeded, break out of the loop
                            break

                        except Exception as e:
                            #print(f"Error: {str(e)}")
                                
                            if retry_count < self.com_max_retries - 1:
                                # Sleep before the next retry
                                time.sleep(self.com_retry_delay)
                                continue  # Retry the operation
                            else:
                                print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
        # Save the changes to the MPP file (optional)
        app.FileSave()

        # Close the project and release resources
        app.FileClose(1)  # 1 indicates to save changes when closing
        app.Quit()
        print("Finished: Finding and deleting MS project tasks that no longer exist in the structure")
    
    
    def getCorrectUniqueIDsFast(self):
        print("Standby: Fetching Unique IDs from MS Project")
        print("NOTE: This is slow on purpose because com32 is slow and error prone.  Please be patient.")
        # Create an instance of Microsoft Project
        try:
            app = win32com.client.Dispatch("MSProject.Application")
        except Exception as e:
            print(f"Error: {str(e)}")

        # Open the MPP file
        mpp_file_path = self.projectFilename
        app.FileOpen(mpp_file_path)
        time.sleep(5)
        # Get a reference to the project
        # Check if there is an active project
        if app.ActiveProject is not None:
            project = app.ActiveProject
            # Now you can work with the active project
        else:
            print("No active project found.")
        
        # Get a reference to the tasks in the project
        tasks = project.Tasks
        
        selectorToID = {}
        
        #populate a dictionary that mates MS project tasks to their ID in order to update the main dictionary
        for task in tasks:
            taskTextTen = ""
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    taskTextTen = str(task.Text10)

                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
                        
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
            if taskTextTen != "":
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        selectorToID[task.Text10] = task.ID

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                            
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
            else:
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        task.Text10 = task.Name

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                            
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        selectorToID[task.Text10] = task.ID

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                            
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
        
        #populates all of the unique IDs of the tasks that already exist in the project
        for selectorName, task_data in self.data_dict.items():
            if selectorName in selectorToID:
                task_data["Unique ID"] = selectorToID[selectorName]
                task_data["NewTask"] = "No"
        
        #populate the tasks found in project that are not in the structure
        for selectorName, value in selectorToID.items():
            if selectorName not in self.data_dict:
                self.tasksNotInStruct[selectorName] = value
                
                for task in tasks:
                    #finds the matching task to then update its info in the data dict
                    if task.Text10 == selectorName:
                        # Build the inner dictionary
                        inner_dict = {
                            "IsProjectBuffer": "No",
                            "ProjectTaskNoStruct": "Yes",
                            "NewTask": "No",
                            "Unique ID": task.ID,
                            "Formatter Assigned ID": "",
                            "Unique ID Successors": task.Successors,
                            "Unique ID Predecessors": task.Predecessors,
                            "Text11": task.Text11,
                            "Text12": task.Text12,
                            "Text13": task.Text13,
                            "Name": task.Name,
                            "Text14": task.Text14,
                            "Text15": task.Text15,
                            "Text16": task.Text16,
                            "Resource Names": task.ResourceNames,
                            "Text17": task.Text17,
                            "Text18": task.Text18,
                            "Text19": task.Text19,
                            "Text20": task.Text20,
                            "Duration": self.convertMinutesToDays(task.Duration),
                            "Low-Risk Dur": self.convertMinutesToDays(task.Duration1),
                            "Text21": task.Text21,
                            "percentcomplete": str(task.PercentComplete)
                        }
                        # Add the inner dictionary to the main dictionary
                        self.data_dict[selectorName] = inner_dict
                    if task.Name == selectorName:
                        # Build the inner dictionary
                        inner_dict = {
                            "IsProjectBuffer": "Yes",
                            "ProjectTaskNoStruct": "No",
                            "NewTask": "No",
                            "Unique ID": task.ID,
                            "Formatter Assigned ID": "",
                            "Unique ID Successors": task.Successors,
                            "Unique ID Predecessors": task.Predecessors,
                            "Text11": task.Text11,
                            "Text12": task.Text12,
                            "Text13": task.Text13,
                            "Name": task.Name,
                            "Text14": task.Text14,
                            "Text15": task.Text15,
                            "Text16": task.Text16,
                            "Resource Names": task.ResourceNames,
                            "Text17": task.Text17,
                            "Text18": task.Text18,
                            "Text19": task.Text19,
                            "Text20": task.Text20,
                            "Duration": self.convertMinutesToDays(task.Duration),
                            "Low-Risk Dur": self.convertMinutesToDays(task.Duration1),
                            "Text21": task.Text21,
                            "percentcomplete": str(task.PercentComplete)
                        }
                        # Add the inner dictionary to the main dictionary
                        self.data_dict[selectorName] = inner_dict
            
        # Save the changes to the MPP file (optional)
        app.FileSave()
        
        app.FileClose(1)  # 1 indicates to save changes when closing
        app.Quit()
        print("Finished: Fetching Unique IDs from MS Project")
    
    def convertMinutesToDays(self, minutes):
        hours = minutes // 60
        days = hours / 8
        daysString = str(days)
        daysString = daysString.split('.')[0]
        completeString = daysString + " days"
        return daysString
    
    def addNewTasksToProject(self):
        print("Standby: Adding New Tasks to MS Project")
        print("NOTE: This is slow on purpose because com32 is slow and error prone.  Please be patient.")
        # Create an instance of Microsoft Project
        try:
            app = win32com.client.Dispatch("MSProject.Application")
        except Exception as e:
            print(f"Error: {str(e)}")

        # Open the MPP file
        mpp_file_path = self.projectFilename
        app.FileOpen(mpp_file_path)
        time.sleep(5)
        # Get a reference to the project
        # Check if there is an active project
        if app.ActiveProject is not None:
            project = app.ActiveProject
            # Now you can work with the active project
        else:
            print("No active project found.")

        # Get a reference to the tasks in the project
        tasks = project.Tasks
        
        # Find the maximum existing Unique ID
        max_unique_id = max([task.UniqueID for task in tasks])
        numberOfNewTasks = 0

        for task_name, task_data in self.data_dict.items():
            if task_data["NewTask"] == "Yes":
                print("Adding the following task: " + task_data["Name"])
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task = tasks.Add()

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text10 = task_name

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Name = task_data["Name"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text11 = task_data["Text11"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text12 = task_data["Text12"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                        
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text13 = task_data["Text13"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text14 = task_data["Text14"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                        
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text15 = task_data["Text15"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                        
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text16 = task_data["Text16"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.ResourceNames = task_data["Resource Names"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text17 = task_data["Text17"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text18 = task_data["Text18"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text19 = task_data["Text19"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text20 = task_data["Text20"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Duration = task_data["Duration"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Duration1 = task_data["Low-Risk Dur"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.Text21 = task_data["Text21"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                for retry_count in range(self.com_max_retries):
                    try:
                        # Open the Microsoft Project file
                        new_task.PercentComplete = task_data["percentcomplete"]

                        # If the operation succeeded, break out of the loop
                        break

                    except Exception as e:
                        #print(f"Error: {str(e)}")
                        
                        if retry_count < self.com_max_retries - 1:
                            # Sleep before the next retry
                            time.sleep(self.com_retry_delay)
                            continue  # Retry the operation
                        else:
                            print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
                max_unique_id += 1
                #new_task.ID = max_unique_id
                task_data["Unique ID"] = str(new_task.ID)
                numberOfNewTasks += 1
                
        print(str(numberOfNewTasks) + " new tasks successfully added to the MS project file")
        # Save the changes to the MPP file (optional)
        app.FileSave()
        
        app.FileClose(1)  # 1 indicates to save changes when closing
        app.Quit()
        
    def correctPredecessorSuccessorLogic(self):
        #Note this only updates it in the data dict and not in the project file until it runs the updatePredecessorandSuccessorLogic method
        print("Standby: Updating Successor and Predecessor Logic")
        formatter_to_unique_mapping = {}
        for key, value in self.data_dict.items():
            formatter_id = value['Formatter Assigned ID']
            unique_id = value['Unique ID']
            formatter_to_unique_mapping[formatter_id] = unique_id

        # Step 2: Update Successor and Predecessor values in the original dictionary
        for key, value in self.data_dict.items():
            #since the tasks not in the struct are already in project they do not need their successor/predecessor logic re-mapped
            if value['Formatter Assigned ID'] != "":
                successor_str = value['Unique ID Successors']
                if successor_str != "":
                    formatter_ids_successors = [int(x.strip()) for x in successor_str.split(',')]
                    unique_ids_successors = [formatter_to_unique_mapping.get(str(formatter_id), "Broken") for formatter_id in formatter_ids_successors]
                    value['Unique ID Successors'] = ','.join(map(str, unique_ids_successors))
                
                predecessor_str = value['Unique ID Predecessors']
                if predecessor_str != "":
                    formatter_ids_predecessors = [int(x.strip()) for x in predecessor_str.split(',')]
                    unique_ids_predecessors = [formatter_to_unique_mapping.get(str(formatter_id), "Broken") for formatter_id in formatter_ids_predecessors]
                    value['Unique ID Predecessors'] = ','.join(map(str, unique_ids_predecessors))
        
        print("Finished: Updating Successor and Predecessor Logic")
                
    def exportDataDictToJSON(self, fileName):
        print("Standby: Exporting Dictionary to JSON")
        # Export the dictionary to a JSON file
        with open(fileName, 'w') as json_file:
            json.dump(self.data_dict, json_file, indent=4)
        print("Finished: Exporting Dictionary to JSON")
    
    def setTasksToManuallyvsAutomaticallyScheduled(self, manualFlag):
        if(manualFlag):
            print("Standby: Changing all tasks to Manually scheduled")
            print("NOTE: This is slow on purpose because com32 is slow and error prone.  Please be patient.")
        else:
            print("Standby: Changing all tasks to Automatically scheduled")
            print("NOTE: This is slow on purpose because com32 is slow and error prone.  Please be patient.")
        
        # Create an instance of Microsoft Project
        try:
            app = win32com.client.Dispatch("MSProject.Application")
        except Exception as e:
            print(f"Error: {str(e)}")

        # Open the MPP file
        mpp_file_path = self.projectFilename
        print("MPP File Path:", mpp_file_path)  #AC: Added for debugging
        app.FileOpen(mpp_file_path)
        time.sleep(5)
        # Get a reference to the project
        # Check if there is an active project
        if app.ActiveProject is not None:
            project = app.ActiveProject
            # Now you can work with the active project
        else:
            print("No active project found.")

        # Get a reference to the tasks in the project
        tasks = project.Tasks
        
        #to update successor and predecessor without pissing off MS project we must make all tasks manually scheduled and then wipe out all existing successor and predecessors then iterate through and update
        for task in tasks:
            # Set scheduling mode to Manual
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    task.Manual = manualFlag

                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
                        
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
        # Save the changes to the MPP file
        app.FileSave()

        # Close the project and release resources
        app.FileClose(1)  # 1 indicates to save changes when closing
        app.Quit()
        if(manualFlag):
            print("Finished: Changing all tasks to Manually scheduled")
        else:
            print("Finished: Changing all tasks to Automatically scheduled")
    
    def updatePredecessorandSuccessorLogic(self):
        print("Standby: Updating the Predecessor and Successor logic for all tasks")
        print("NOTE: This is slow on purpose because com32 is slow and error prone.  Please be patient.")
        # Create an instance of Microsoft Project
        try:
            app = win32com.client.Dispatch("MSProject.Application")
        except Exception as e:
            print(f"Error: {str(e)}")

        # Open the MPP file
        mpp_file_path = self.projectFilename
        app.FileOpen(mpp_file_path)
        time.sleep(5)
        # Get a reference to the project
        # Check if there is an active project
        if app.ActiveProject is not None:
            project = app.ActiveProject
            # Now you can work with the active project
        else:
            print("No active project found.")

        # Get a reference to the tasks in the project
        tasks = project.Tasks
        
        for task in tasks:
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    task.Successors = self.data_dict[task.Text10]["Unique ID Successors"]

                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
                
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
            
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    task.Predecessors = self.data_dict[task.Text10]["Unique ID Predecessors"]

                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
                
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
        # Save the changes to the MPP file (optional)
        app.FileSave()

        # Close the project and release resources
        app.FileClose(1)  # 1 indicates to save changes when closing
        app.Quit()
        print("Finished: Updating the Predecessor and Successor logic for all tasks")
    
    def updateDurationsAndCompletionPercent(self):
        print("Standby: Updating the Low Risk Duration, Focused Duration, Duration Remaining, and Percent Complete of all tasks.  Also updating ART because managers love to move teams ARTs just because")
        print("NOTE: This is slow on purpose because com32 is slow and error prone.  Please be patient.")
        # Create an instance of Microsoft Project
        try:
            app = win32com.client.Dispatch("MSProject.Application")
        except Exception as e:
            print(f"Error: {str(e)}")

        # Open the MPP file
        mpp_file_path = self.projectFilename
        app.FileOpen(mpp_file_path)
        time.sleep(5)
        # Get a reference to the project
        # Check if there is an active project
        if app.ActiveProject is not None:
            project = app.ActiveProject
            # Now you can work with the active project
        else:
            print("No active project found.")

        # Get a reference to the tasks in the project
        tasks = project.Tasks
        
        for task in tasks:
            taskSelector = ""
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    taskSelector = str(task.Text10)
                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
                    
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
            
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    assignments = task.Assignments
                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
                    
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
            
            #This code factors in when you force MS project to use more than 100% of the resource group and ensures you do not overwrite the durations undoing that.
            resource_units = 1
            for retry_count in range(self.com_max_retries):
                try:
                    if assignments is not None:
                        for assignment in assignments:
                            if assignment.Task.ID == task.id:
                                if assignment.Units > 1:
                                    resource_units = assignment.Units
                except Exception as e:
                    #print(f"Error: {str(e)}")
            
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
            for retry_count in range(self.com_max_retries):
                try:
                    parts = str(self.data_dict[task.Text10]["Duration"]).split()
                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
            
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
            
            tempDuration = float(parts[0]) / resource_units
            if len(parts) == 2:
                duration= str(tempDuration) + " " + parts[1]
            else:
                duration= str(tempDuration) + " days"
            
            for retry_count in range(self.com_max_retries):
                try:
                    parts = str(self.data_dict[task.Text10]["Low-Risk Dur"]).split()
                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
            
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
            
            tempDurationOne = float(parts[0]) / resource_units
            if len(parts) == 2:
                durationOne= str(tempDurationOne) + " " + parts[1]
            else:
                durationOne= str(tempDurationOne) + " days"
            if(self.data_dict[task.Text10]["Text21"] != ""):
                parts = str(self.data_dict[task.Text10]["Text21"]).split()
                tempTextTwentyOne = float(parts[0]) / resource_units
                textTwentyOne= str(tempTextTwentyOne) + " " + parts[1]
            
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    task.Duration = duration

                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
            
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                    
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    task.Duration1 = durationOne

                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
            
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                    
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    task.Text21 = textTwentyOne

                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
            
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
        
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    task.PercentComplete = self.data_dict[task.Text10]["percentcomplete"]
                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
            
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                
            for retry_count in range(self.com_max_retries):
                try:
                    # Open the Microsoft Project file
                    task.Text16= self.data_dict[task.Text10]["Text16"]
                    # If the operation succeeded, break out of the loop
                    break

                except Exception as e:
                    #print(f"Error: {str(e)}")
            
                    if retry_count < self.com_max_retries - 1:
                        # Sleep before the next retry
                        time.sleep(self.com_retry_delay)
                        continue  # Retry the operation
                    else:
                        print("RETRIES DIDNT WORK PLEASE CONSIDER INCREASING RETRIES TO MAKE IT LESS ERROR PRONE IN WIN32")
                            
        # Save the changes to the MPP file (optional)
        app.FileSave()

        # Close the project and release resources
        app.FileClose(1)  # 1 indicates to save changes when closing
        app.Quit()
        print("Finished: Updating the Low Risk Duration, Focused Duration, Duration Remaining, and Percent Complete of all tasks")  
    
    def debugOnlyBuildDictionaryFromJSON(self):
        try:
            json_file_path = "GroundTruth.json"
            with open(json_file_path, 'r') as json_file:
                self.data_dict = json.load(json_file)
        except Exception as e:
            print(f"Error reading the JSON file: {e}")
        
    def errorCheckTheFormatterOutput(self):
        print("Standby: Error Checking the Excel Output")
        try:
            workbook = openpyxl.load_workbook(self.formattedExcelForImport)
            sheet = workbook.active
            errorFound = False

            # Assuming the data starts from row 2
            for rowPresent in sheet.iter_rows(min_row=2, values_only=True):
                if "TAP-" in str(rowPresent[2]):
                    print("TAP ticket found in Successor Column of the formatted JIRA Export.")
                    print("This is caused when a ticket is marked as a successor and is not actually in the structure.")
                    print("Script will terminate so you can go investigate the issue and likely pull that feature into the structure and start over.")
                    print("The offending ticket(s) are: " + str(rowPresent[2]))
                    errorFound=True
                if "TAP-" in str(rowPresent[3]):
                    print("TAP ticket found in Predecessor Column of the formatted JIRA Export.")
                    print("This is caused when a ticket is marked as a predecessor and is not actually in the structure.")
                    print("Script will terminate so you can go investigate the issue and likely pull that feature into the structure and start over.")
                    print("The offending ticket(s) are: " + str(rowPresent[2]))
                    errorFound=True
                if str(rowPresent[9]) == "Feature" and str(rowPresent[2]) == "" and "Milestone" not in str(rowPresent[8]) and "Prochain" not in str(rowPresent[8]):
                    print("TAP feature found that has no successor and it is not a project buffer feature.")
                    print("This will cause scheduling problems because it creates a fake endpoint in the schedule.")
                    print("Please go fix the tickets successor relationship in JIRA and start over.")
                    print("The offending ticket is: " + str(rowPresent[1]))
                    errorFound=True
                #Add checks for duplicated lines indicating duplicate features in the structure because this makes this automation super pissed
            workbook.save(self.formattedExcelForImport)
            workbook.close()
            if errorFound:
                print("Exiting because of the errors found above in the excel.  Correct these and try again")
                sys.exit(1)
                    
        except Exception as e:
            print(f"Error reading the Excel file: {e}")
        print("Excel generated from formatter is good to go!")
        
    def exportSummaryReport(self):
        print("Standby: Creating and exporting a schedule update summary report")
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Added_Tasks"
            rowIndex = 4
            countNewTasks = 0
            header_fill = PatternFill(start_color='595959', end_color='595959', fill_type='solid')
            header_font = Font(bold=True, color=WHITE)
            sheet.cell(row=3, column=1, value="MS Project Unique ID")
            sheet['A3'].fill = header_fill
            sheet['A3'].font = header_font
            sheet['A3'].alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=3, column=2, value="Name")
            sheet['B3'].fill = header_fill
            sheet['B3'].font = header_font
            sheet['B3'].alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=3, column=3, value="Selector")
            sheet['C3'].fill = header_fill
            sheet['C3'].font = header_font
            sheet['C3'].alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=3, column=4, value="Focused Duration")
            sheet['D3'].fill = header_fill
            sheet['D3'].font = header_font
            sheet['D3'].alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=3, column=5, value="Low Risk Duration")
            sheet['E3'].fill = header_fill
            sheet['E3'].font = header_font
            sheet['E3'].alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=3, column=6, value="Remaining Duration")
            sheet['F3'].fill = header_fill
            sheet['F3'].font = header_font
            sheet['F3'].alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=3, column=7, value="Type")
            sheet['G3'].fill = header_fill
            sheet['G3'].font = header_font
            sheet['G3'].alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=3, column=8, value="Resource")
            sheet['H3'].fill = header_fill
            sheet['H3'].font = header_font
            sheet['H3'].alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=3, column=9, value="Percent Complete")
            sheet['I3'].fill = header_fill
            sheet['I3'].font = header_font
            sheet['I3'].alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=3, column=10, value="Status")
            sheet['J3'].fill = header_fill
            sheet['J3'].font = header_font
            sheet['J3'].alignment = Alignment(horizontal='center', vertical='center')
            
            for task_name, task_data in self.data_dict.items():
                if task_data["NewTask"] == "Yes":
                    sheet.cell(row=rowIndex, column=1, value=task_data["Unique ID"])
                    sheet.cell(row=rowIndex, column=2, value=task_data["Name"])
                    sheet.cell(row=rowIndex, column=3, value=task_name)
                    sheet.cell(row=rowIndex, column=4, value=task_data["Duration"])
                    sheet.cell(row=rowIndex, column=5, value=task_data["Low-Risk Dur"])
                    sheet.cell(row=rowIndex, column=6, value=task_data["Text21"])
                    sheet.cell(row=rowIndex, column=7, value=task_data["Text14"])
                    sheet.cell(row=rowIndex, column=8, value=task_data["Resource Names"])
                    sheet.cell(row=rowIndex, column=9, value=task_data["percentcomplete"])
                    sheet.cell(row=rowIndex, column=10, value=task_data["Text15"])
                    countNewTasks += 1
                    rowIndex += 1
            sheet.cell(row=1, column=1, value= str(countNewTasks) + " New Tasks Added To The Project and Listed Below")
            sheet.auto_filter.ref = 'A3:J3'
            
            #auto adjust the column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass
                adjusted_width = (max_length)  # Add some extra space
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            new_sheet = workbook.create_sheet(title='Wendy Fix Tickets')
            new_sheet.cell(row=3, column=1, value="Selector")
            header_fill = PatternFill(start_color='595959', end_color='595959', fill_type='solid')
            header_font = Font(bold=True, color=WHITE)
            new_sheet.cell(row=3, column=1, value="MS Project Unique ID")
            new_sheet['A3'].fill = header_fill
            new_sheet['A3'].font = header_font
            new_sheet['A3'].alignment = Alignment(horizontal='center', vertical='center')
            new_sheet.cell(row=3, column=2, value="Name")
            new_sheet['B3'].fill = header_fill
            new_sheet['B3'].font = header_font
            new_sheet['B3'].alignment = Alignment(horizontal='center', vertical='center')
            new_sheet.cell(row=3, column=3, value="Selector")
            new_sheet['C3'].fill = header_fill
            new_sheet['C3'].font = header_font
            new_sheet['C3'].alignment = Alignment(horizontal='center', vertical='center')
            new_sheet.cell(row=3, column=4, value="Focused Duration")
            new_sheet['D3'].fill = header_fill
            new_sheet['D3'].font = header_font
            new_sheet['D3'].alignment = Alignment(horizontal='center', vertical='center')
            new_sheet.cell(row=3, column=5, value="Low Risk Duration")
            new_sheet['E3'].fill = header_fill
            new_sheet['E3'].font = header_font
            new_sheet['E3'].alignment = Alignment(horizontal='center', vertical='center')
            new_sheet.cell(row=3, column=6, value="Remaining Duration")
            new_sheet['F3'].fill = header_fill
            new_sheet['F3'].font = header_font
            new_sheet['F3'].alignment = Alignment(horizontal='center', vertical='center')
            new_sheet.cell(row=3, column=7, value="Type")
            new_sheet['G3'].fill = header_fill
            new_sheet['G3'].font = header_font
            new_sheet['G3'].alignment = Alignment(horizontal='center', vertical='center')
            new_sheet.cell(row=3, column=8, value="Resource")
            new_sheet['H3'].fill = header_fill
            new_sheet['H3'].font = header_font
            new_sheet['H3'].alignment = Alignment(horizontal='center', vertical='center')
            new_sheet.cell(row=3, column=9, value="Percent Complete")
            new_sheet['I3'].fill = header_fill
            new_sheet['I3'].font = header_font
            new_sheet['I3'].alignment = Alignment(horizontal='center', vertical='center')
            new_sheet.cell(row=3, column=10, value="Status")
            new_sheet['J3'].fill = header_fill
            new_sheet['J3'].font = header_font
            new_sheet['J3'].alignment = Alignment(horizontal='center', vertical='center')
            rowIndex = 4
            countWendyTasks = 0
            for ticketName in self.badTicketFixTicketsMarkedComplete:
                new_sheet.cell(row=rowIndex, column=2, value=ticketName)
                countWendyTasks += 1
                rowIndex += 1
            new_sheet.cell(row=1, column=1, value= str(countWendyTasks) + " Tasks skipped due to the Wendy Fix)")
            new_sheet.auto_filter.ref = 'A3:J3'
            
            #auto adjust the column widths
            for column in new_sheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass
                adjusted_width = (max_length)  # Add some extra space
                new_sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(self.pathToScheduleUpdateSummary)
            workbook.close()

        except Exception as e:
            print(f"Error reading the Excel file: {e}")
        print("Finished: Creating and exporting a schedule update summary report")
    
    def duplicateMPPFile(self):
        print("Standby: Duplicating Original Project File")
        try:
            shutil.copyfile(self.baselineScheduleFilename, self.projectFilename)
        except Exception as e:
            print(f"Error: {str(e)}")
        print("Finished: Duplicating Original Project File")
    
    def killMSProjectIfOpen(self):
        print("Standby: Terminating MS Project From Running")
        for process in psutil.process_iter(attrs=['pid', 'name']):
            if process.info['name'] == "WINPROJ.EXE":
                try:
                    # Terminate (kill) the Excel process
                    psutil.Process(process.info['pid']).terminate()
                except psutil.NoSuchProcess:
                    pass  # Handle the case if the process no longer exists
        print("Finished: Terminating MS Project From Running")
    
    def killExcelIfOpen(self):
        print("Standby: Terminating Excel From Running")
        for process in psutil.process_iter(attrs=['pid', 'name']):
            if process.info['name'] == "EXCEL.EXE":
                try:
                    # Terminate (kill) the Excel process
                    psutil.Process(process.info['pid']).terminate()
                except psutil.NoSuchProcess:
                    pass  # Handle the case if the process no longer exists
        print("Finished: Terminating Excel From Running")        

    # this method basically allows you to feed in an array of tickets you want to mark 100% complete because you know they are bad and not correct
    def applyBadTicketFix(self):
        print("Standby: Applying the Bad Ticket Fix")
        #badTicketFixTickets = ["TAP-87417", "TAP-84549", "TAP-83547", "TAP-88236", "TAP-88387", "TAP-85483", "TAP-88231", "TAP-88386", "TAP-57120", "TAP-84563", "TAP-87440", "TAP-86973", "TAP-85482", "TAP-85484", "TAP-67835", "TAP-67834", "TAP-84546", "TAP-88384", "TAP-87425", "TAP-88385", "TAP-87426", "TAP-95494", "TAP-64261", "TAP-97622"]
        badTicketFixTickets = []
        try:
            workbook = openpyxl.load_workbook(self.formattedExcelForImport)
            sheet = workbook.active

            # Iterate through rows
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Assuming the data starts from row 2
                tapTicketOnlyNoScrum = row[0].split('_')[0]
                if tapTicketOnlyNoScrum in badTicketFixTickets:  # Assuming column A is the first column (index 0)
                    if self.debugFlag:
                        print("Applying Wendy Fix to: " + row[7])
                    self.badTicketFixTicketsMarkedComplete.append(row[7])
                    cell_to_update = f'T{row_index}'  # Assuming you want to update column T
                    sheet[cell_to_update] = 1
        
            workbook.save(self.formattedExcelForImport)
            workbook.close()

        except Exception as e:
            print(f"Error reading the Excel file: {e}")
        print("Finished: Applying Bad Ticket Fix")
    
    def RunTheRoutine(self):
        self.killMSProjectIfOpen()
        self.killExcelIfOpen()
        #self.errorCheckStructureOutput()
        #self.formatTheJIRAExport()
        self.errorCheckTheFormatterOutput()
        self.applyBadTicketFix()
        self.duplicateMPPFile()
        self.setTasksToManuallyvsAutomaticallyScheduled(True) #turns all tasks to manually scheduled except project buffers
        self.buildDictionaryFromFormatterExport()
        if self.debugFlag:
            self.exportDataDictToJSON("Raw.json")
        self.deleteProjectTasksNoLongerInStruct()
        self.getCorrectUniqueIDsFast()
        if self.debugFlag:
            self.exportDataDictToJSON("CorrectedIDs.json")
        self.addNewTasksToProject()
        if self.debugFlag:
            self.exportDataDictToJSON("AnotherOne.json")
        self.correctPredecessorSuccessorLogic()
        if self.debugFlag:
            self.exportDataDictToJSON("CorrectedPredecessorSuccessor.json")
        #self.setTasksToManuallyvsAutomaticallyScheduled(True) #turns all tasks to manually scheduled except project buffers
        self.updatePredecessorandSuccessorLogic()
        self.updateDurationsAndCompletionPercent()
        self.setTasksToManuallyvsAutomaticallyScheduled(False) #turns all tasks to automatically scheduled again
        self.exportSummaryReport()
        print("ALL DONE, Enjoy the updated schedule and hope its on track")